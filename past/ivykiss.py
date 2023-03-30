# %% import os
import logging
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
import requests
import pandas as pd
import json

import os

# get the current working directory
current_dir = os.getcwd()

# Configure logging settings
log_filename = current_dir+"my_log_file.log"
logging.basicConfig(filename=log_filename, filemode="a", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# %%
with open(r'C:\Users\KISS Admin\Desktop\IVYENT_DH\data.json', 'r') as f:
    data = json.load(f)

driver = webdriver.Chrome()
driver.maximize_window()  # 창 최대화

driver.get(data['ivykiss_url'])

from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from dateutil import relativedelta

# %%
# ID와 비밀번호 가져오기
server = data['server']
database = data['database']
username = data['username']
password = data['password']

connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")
logging.info("Connection Established:")

df_mtrl= pd.read_sql("""WITH T1 as (
( SELECT material FROM [ivy.mm.dim.fact_poasn] WHERE act_date >=getdate() GROUP BY material )
UNION 
(SELECT material from [ivy.mm.dim.mrp01] WHERE total_stock>0 GROUP BY material)
)
SELECT T1.* FROM T1
LEFT JOIN sku_image T2 on T1.material=T2.Material
LEFT JOIN [ivy.mm.dim.mtrl] T3 on T1.material = T3.material
WHERE link is null
and T3.ms in ('01','41','91')
ORDER BY T1.material
""", con=engine
)
df_mtrl.head()

# %%
# a=input("After setting and login to ivykiss, press any key to continue")
time.sleep(5)   
lth=len(df_mtrl)
#   %%
login=driver.find_element(By.CLASS_NAME,"input_text")
login.click()
login.send_keys(data['ivykiss_id'])

pw=driver.find_elements(By.CLASS_NAME,"input_text")[1]
pw.click()
pw.send_keys(data['ivykiss_pw'])
driver.find_element(By.CLASS_NAME,"button").click() # click login button

time.sleep(5)   

img_list=[]
cnt=0
for index, row in df_mtrl.iterrows():
    # if index>3415:
    try:
        # row.material="RMPF15"
        url = data['search_url2']+ row.material
        
        driver.get(url)
        time.sleep(1)   
# Locate the 'img-gallery' div element
        items=driver.find_elements(By.CLASS_NAME, "product-item-info")
        inlist=[row.material]
        for item in items:
            item_name= item.find_element(By.CLASS_NAME, "product-sku").get_attribute('innerHTML')
            # item=items[0]
            if row.material== item_name:
                
                img_gallery = item.find_element(By.CLASS_NAME, "img-gallery")

                # Find all the 'img' and 'a' elements inside the 'img-gallery' div element
                img_links = img_gallery.find_elements(By.CLASS_NAME,'img-link')

                # Loop through each 'img' and 'a' element and get the 'src' or 'href' attribute value
                for img_link in img_links:
                    a_href = img_link.get_attribute("href")
                    con1= 'front' in a_href
                    con2= 'FRONT' in a_href
                    con3= 'Front' in a_href
                    if con1|con2|con3 :
                        inlist.append(a_href)                    
        img_list.append(inlist)
        if len(inlist)>1:
            cnt=cnt+1
            print(str(cnt)+" images found in "+ str(len(img_list))+" " + str(lth) + " total materials")
            print(inlist)
            logging.info(f"{cnt} images found in {len(img_list)} {lth} total materials")
            logging.info(inlist)
    except:
        print(f"Error processing material: {row.material}")
        logging.error(f"Error processing material: {row.material}")

img_list= pd.DataFrame(img_list)
img_list.to_csv('ivykiss_image.csv')

driver.quit()

