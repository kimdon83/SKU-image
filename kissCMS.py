import os
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
import requests
import pandas as pd
import json

with open('data.json', 'r') as f:
    data = json.load(f)

driver = webdriver.Chrome()
driver.maximize_window()  # 창 최대화

driver.get(data['okta_url'])

from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from dateutil import relativedelta

# %%
# ID와 비밀번호 가져오기
server = data['server']
database = data['database']
username = data['username']
password = data['password']

connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")

df_mtrl= pd.read_sql("""( SELECT material FROM [ivy.mm.dim.fact_poasn] WHERE act_date >=getdate() GROUP BY material )
UNION 
(SELECT material from [ivy.mm.dim.mrp01] WHERE total_stock>0 GROUP BY material)
ORDER BY material""", con=engine
)
df_mtrl.head()
# %%
a=input("you have to login to okta before going to cms")

img_list=[]
for index, row in df_mtrl.iterrows():
    # if index>3415:
    try:
        url = data['search_url']+ row.material+"&order_by=modified"
        
        driver.get(url)
        time.sleep(1)   
        Images=driver.find_elements(By.CLASS_NAME,"ImageWrapperLarge")
        inlist=[row.material]
        for index0 in range(len(Images)):
            con1='Front' in Images[index0].get_attribute('title')
            con2='front' in Images[index0].get_attribute('title')
            con3= row.material in Images[index0].get_attribute('title')
            if ((con1 | con2) &con3):
                html = Images[index0].get_attribute('innerHTML')
                soup = BeautifulSoup(html, 'html.parser')
                img_tag = soup.find('img')
                img_url = img_tag.get('src')
                inlist.append(img_url)
        img_list.append(inlist)
    except:
        print(row.material)

img_list= pd.DataFrame(img_list)
img_list.to_csv('img_test3.csv')

driver.quit()