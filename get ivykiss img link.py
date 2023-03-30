# %% import os
import logging
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from bs4 import BeautifulSoup
import requests
import pandas as pd
import json

import os

start = time.time()

# get the current working directory
current_dir = os.getcwd()

# Configure logging settings
# log_filename = current_dir+"my_log_file.log"
# logging.basicConfig(filename=log_filename, filemode="a", level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# %%
with open(r'C:\Users\KISS Admin\Desktop\IVYENT_DH\data.json', 'r') as f:
    data = json.load(f)

driver = webdriver.Chrome()
driver.maximize_window()  # 창 최대화

driver.get(data['ivykiss_url'])

from sqlalchemy import create_engine

from sqlalchemy.engine import URL
from datetime import datetime
from dateutil.relativedelta import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from dateutil import relativedelta

# %%
today = datetime.today().date()
today_str = today.strftime('%Y-%m-%d')

# %%
# ID와 비밀번호 가져오기
server = data['server']
database = data['database']
username = data['username']
password = data['password']

connection_string = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+ password
connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": connection_string})
engine = create_engine(connection_url)
print("Connection Established:")

df_mtrl= pd.read_sql("""WITH invT as (
    SELECT material, sum(total_stock) as total_stock FROM [ivy.mm.dim.mrp01]
    WHERE total_stock>0
    GROUP BY material
), msmtrl as (
    SELECT T1.material, T1.ms 

    FROM [ivy.mm.dim.mtrl] T1
    LEFT JOIN sku_image T2 on T1.material=T2.Material
    LEFT JOIN invT on T1.material = invT.material
    WHERE T1.ms in ('01','03','D1','N1')
    and T1.mg <> 'pp'
    and T2.Link is null
    and invT.material is not null
)
SELECT material FROM msmtrl ORDER BY material
""", con=engine
)

# df_mtrl= pd.read_csv(current_dir+'\ivykiss_no front to get 0324.csv')
df_mtrl.head()
# %%
time.sleep(5)   
lth=len(df_mtrl)
#   %%
login=driver.find_element(By.CLASS_NAME,"input_text")
login.click()
login.send_keys(data['ivykiss_id'])

pw=driver.find_elements(By.CLASS_NAME,"input_text")[1]
pw.click()
pw.send_keys(data['ivykiss_pw'])
driver.find_element(By.CLASS_NAME,"button").click() # click login button

time.sleep(5)   

img_list=[]
img_notfront=[]
cnt=0
for index, row in df_mtrl.iterrows():
    try:
        url = data['search_url2']+ row.material
        
        driver.get(url)
        time.sleep(1)   
        items=driver.find_elements(By.CLASS_NAME, "product-item-info")
        inlist=[row.material]
        inlist_nf=[row.material]
        for item in items:
            item_name= item.find_element(By.CLASS_NAME, "product-sku").get_attribute('innerHTML')
            if row.material== item_name:
                
                img_gallery = item.find_element(By.CLASS_NAME, "img-gallery")

                img_links = img_gallery.find_elements(By.CLASS_NAME,'img-link')

                for img_link in img_links:
                    a_href = img_link.get_attribute("href")
                    con1= 'front' in a_href
                    con2= 'FRONT' in a_href
                    con3= 'Front' in a_href
                    if con1|con2|con3==True:
                        inlist.append(a_href)
                    else:
                        inlist_nf.append(a_href)                   
        img_list.append(inlist)
        if len(inlist)==1:
            img_notfront.append(inlist_nf)
        if len(inlist)>1:
            cnt=cnt+1
            print(str(cnt)+" images found in "+ str(index)+" " + str(lth) + " total materials")
            print(inlist)
        elif len(inlist_nf)>1:
            cnt=cnt+1
            print(str(cnt)+" images found in "+ str(index)+" " + str(lth) + " total materials")
            print(inlist_nf)
    except:
        print(row)

img_list= pd.DataFrame(img_list)
img_notfront= pd.DataFrame(img_notfront)
img_list.columns = img_list.columns.astype(str)
img_list = img_list.rename(columns={'0': 'material'})
img_notfront.columns = img_notfront.columns.astype(str)
img_notfront = img_notfront.rename(columns={'0': 'material'})

# img_list.to_csv(today_str+'ivykiss_result.csv')
# img_notfront.to_csv(today_str+'ivykiss_image_no front result.csv')

img_list.loc[pd.isnull(img_list["1"])==False].to_csv(today_str+'ivykiss_result.csv')
img_list.loc[pd.isnull(img_list["1"])==False].to_csv('Product Photo.csv',index=False)
img_notfront.loc[pd.isnull(img_notfront["1"])==False].to_csv(today_str+'ivykiss_image_no front result.csv')
img_notfront.loc[pd.isnull(img_notfront["1"])==False].to_csv('front_check_list.csv', index=False)
end = time.time()
logging.info(end-start)
print(end-start)
driver.quit()

